#!/usr/bin/env python3
"""
Setup script to create sample data and run tests.

Usage:
    python setup_and_test.py          # Create data and run tests
    python setup_and_test.py --only-data   # Only create data
    python setup_and_test.py --only-test   # Only run tests
"""

import argparse
import subprocess
import sys
from pathlib import Path


def create_sample_data():
    """Create the sample Excel file."""
    print("=" * 60)
    print("CREATING SAMPLE DATA")
    print("=" * 60)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Installing openpyxl...")
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

    output_path = Path("data/companies.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    whitelist_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    blacklist_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")

    # Whitelist
    ws_whitelist = workbook.active
    ws_whitelist.title = "Whitelist"

    headers = ["Company Name", "Category", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws_whitelist.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = whitelist_fill

    whitelist_data = [
        ("Siemens AG", "Technology", "Major German technology corporation"),
        ("BMW Group", "Automotive", "Luxury vehicle manufacturer"),
        ("SAP SE", "Software", "Enterprise software solutions"),
        ("Volkswagen AG", "Automotive", "World's largest automaker"),
        ("Mercedes-Benz AG", "Automotive", "Premium vehicle manufacturer"),
        ("Bosch GmbH", "Technology", "Engineering and electronics"),
        ("BASF SE", "Chemical", "Largest chemical producer"),
        ("Allianz SE", "Insurance", "Financial services"),
        ("Deutsche Bank AG", "Finance", "Investment banking"),
        ("Deutsche Telekom AG", "Telecommunications", "Telecom provider"),
        ("Airbus SE", "Aerospace", "Aircraft manufacturer"),
        ("Continental AG", "Automotive", "Tire and auto parts"),
        ("Infineon Technologies AG", "Semiconductors", "Chip manufacturer"),
        ("Henkel AG", "Consumer Goods", "Consumer products"),
        ("ThyssenKrupp AG", "Industrial", "Steel and services"),
        ("Fresenius SE", "Healthcare", "Healthcare company"),
        ("Deutsche Post DHL", "Logistics", "Postal and logistics"),
        ("E.ON SE", "Energy", "Energy company"),
        ("RWE AG", "Energy", "Electric utilities"),
        ("Munich Re", "Insurance", "Reinsurance"),
        ("Adidas AG", "Consumer Goods", "Sporting goods"),
        ("Bayer AG", "Pharmaceutical", "Life sciences"),
        ("Porsche AG", "Automotive", "Sports cars"),
        ("Audi AG", "Automotive", "Premium automobiles"),
        ("ZF Friedrichshafen AG", "Automotive", "Auto supplier"),
        ("Carl Zeiss AG", "Optics", "Optoelectronics"),
        ("Merck KGaA", "Pharmaceutical", "Science and technology"),
        ("Kuka AG", "Robotics", "Industrial robots"),
        ("Trumpf GmbH", "Manufacturing", "Machine tools"),
        ("Zalando SE", "E-Commerce", "Online fashion"),
    ]

    for row, data in enumerate(whitelist_data, start=2):
        for col, value in enumerate(data, start=1):
            ws_whitelist.cell(row=row, column=col, value=value)

    ws_whitelist.column_dimensions["A"].width = 35
    ws_whitelist.column_dimensions["B"].width = 20
    ws_whitelist.column_dimensions["C"].width = 40

    # Blacklist
    ws_blacklist = workbook.create_sheet("Blacklist")

    for col, header in enumerate(headers, start=1):
        cell = ws_blacklist.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = blacklist_fill

    blacklist_data = [
        ("Fake Company GmbH", "Unknown", "Known fraudulent operation"),
        ("Scam Industries Ltd", "Unknown", "Multiple fraud reports"),
        ("Nicht Existiert AG", "Unknown", "Company does not exist"),
        ("Betrug & Partner KG", "Unknown", "Deceptive practices"),
        ("Unseriös Consulting", "Consulting", "Unreliable"),
        ("Abzocke Services GmbH", "Services", "Exploitation reports"),
        ("Phantomfirma SE", "Unknown", "Shell company"),
        ("Scheinfirma International", "Unknown", "No legitimate ops"),
    ]

    for row, data in enumerate(blacklist_data, start=2):
        for col, value in enumerate(data, start=1):
            ws_blacklist.cell(row=row, column=col, value=value)

    ws_blacklist.column_dimensions["A"].width = 35
    ws_blacklist.column_dimensions["B"].width = 20
    ws_blacklist.column_dimensions["C"].width = 40

    workbook.save(output_path)

    print(f"\n✓ Created: {output_path}")
    print(f"  Whitelist: {len(whitelist_data)} companies")
    print(f"  Blacklist: {len(blacklist_data)} companies")

    return str(output_path)


def install_package():
    """Install the package in development mode."""
    print("\n" + "=" * 60)
    print("INSTALLING PACKAGE")
    print("=" * 60)

    result = subprocess.run(
        [sys.executable, "-m", "pip", "install", "-e", "."],
        capture_output=True,
        text=True,
    )

    if result.returncode != 0:
        print(f"Warning: pip install failed: {result.stderr}")
        return False

    print("✓ Package installed")
    return True


def run_quantification(excel_file: str):
    """Run quantification tests."""
    print("\n" + "=" * 60)
    print("RUNNING QUANTIFICATION TESTS")
    print("=" * 60)

    result = subprocess.run(
        [sys.executable, "run_quantification.py", excel_file],
        capture_output=False,
    )

    return result.returncode == 0


def run_pytest():
    """Run pytest on test_quantification.py."""
    print("\n" + "=" * 60)
    print("RUNNING PYTEST")
    print("=" * 60)

    result = subprocess.run(
        [sys.executable, "-m", "pytest", "tests/test_quantification.py", "-v", "--tb=short"],
        capture_output=False,
    )

    return result.returncode == 0


def main():
    parser = argparse.ArgumentParser(description="Setup and test company lookup")
    parser.add_argument("--only-data", action="store_true", help="Only create data")
    parser.add_argument("--only-test", action="store_true", help="Only run tests")
    parser.add_argument("--pytest", action="store_true", help="Run pytest instead of quantification")

    args = parser.parse_args()

    # Change to script directory
    import os
    os.chdir(Path(__file__).parent)

    excel_file = "data/companies.xlsx"

    if not args.only_test:
        excel_file = create_sample_data()
        install_package()

    if not args.only_data:
        if args.pytest:
            success = run_pytest()
        else:
            success = run_quantification(excel_file)

        if success:
            print("\n" + "=" * 60)
            print("✓ ALL TESTS PASSED")
            print("=" * 60)
        else:
            print("\n" + "=" * 60)
            print("✗ SOME TESTS FAILED")
            print("=" * 60)
            sys.exit(1)


if __name__ == "__main__":
    main()
