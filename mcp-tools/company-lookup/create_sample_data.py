#!/usr/bin/env python3
"""Create sample Excel file with company data."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def create_sample_excel(output_path: str = "data/companies.xlsx") -> None:
    """Create a sample Excel file with company whitelist/blacklist data."""
    workbook = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    whitelist_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    blacklist_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Create Whitelist sheet
    ws_whitelist = workbook.active
    ws_whitelist.title = "Whitelist"

    # Headers
    headers = ["Company Name", "Category", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws_whitelist.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = whitelist_fill
        cell.alignment = header_alignment

    # Whitelist companies (realistic German companies)
    whitelist_data = [
        ("Siemens AG", "Technology", "Major German technology corporation"),
        ("BMW Group", "Automotive", "Luxury vehicle manufacturer"),
        ("SAP SE", "Software", "Enterprise software solutions"),
        ("Volkswagen AG", "Automotive", "World's largest automaker by revenue"),
        ("Mercedes-Benz AG", "Automotive", "Premium vehicle manufacturer"),
        ("Bosch GmbH", "Technology", "Engineering and electronics company"),
        ("BASF SE", "Chemical", "Largest chemical producer globally"),
        ("Allianz SE", "Insurance", "Financial services and insurance"),
        ("Deutsche Bank AG", "Finance", "Investment banking"),
        ("Deutsche Telekom AG", "Telecommunications", "Telecom provider"),
        ("Airbus SE", "Aerospace", "Commercial aircraft manufacturer"),
        ("Continental AG", "Automotive", "Tire and automotive parts"),
        ("Infineon Technologies AG", "Semiconductors", "Chip manufacturer"),
        ("Henkel AG", "Consumer Goods", "Consumer and industrial products"),
        ("ThyssenKrupp AG", "Industrial", "Steel and industrial services"),
        ("Fresenius SE", "Healthcare", "Healthcare company"),
        ("Deutsche Post DHL", "Logistics", "Postal and logistics services"),
        ("E.ON SE", "Energy", "Energy company"),
        ("RWE AG", "Energy", "Electric utilities"),
        ("Munich Re", "Insurance", "Reinsurance company"),
        ("Adidas AG", "Consumer Goods", "Sporting goods"),
        ("Bayer AG", "Pharmaceutical", "Pharmaceutical and life sciences"),
        ("Porsche AG", "Automotive", "Sports car manufacturer"),
        ("Audi AG", "Automotive", "Premium automobile manufacturer"),
        ("ZF Friedrichshafen AG", "Automotive", "Automotive supplier"),
        ("Carl Zeiss AG", "Optics", "Optics and optoelectronics"),
        ("Merck KGaA", "Pharmaceutical", "Science and technology company"),
        ("HeidelbergCement AG", "Construction", "Building materials"),
        ("Kuka AG", "Robotics", "Industrial robots"),
        ("Trumpf GmbH", "Manufacturing", "Machine tools and lasers"),
    ]

    for row, (name, category, notes) in enumerate(whitelist_data, start=2):
        ws_whitelist.cell(row=row, column=1, value=name)
        ws_whitelist.cell(row=row, column=2, value=category)
        ws_whitelist.cell(row=row, column=3, value=notes)

    # Adjust column widths
    ws_whitelist.column_dimensions["A"].width = 35
    ws_whitelist.column_dimensions["B"].width = 20
    ws_whitelist.column_dimensions["C"].width = 50

    # Create Blacklist sheet
    ws_blacklist = workbook.create_sheet("Blacklist")

    # Headers
    for col, header in enumerate(headers, start=1):
        cell = ws_blacklist.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = blacklist_fill
        cell.alignment = header_alignment

    # Blacklist companies (fictional problematic companies)
    blacklist_data = [
        ("Fake Company GmbH", "Unknown", "Known fraudulent operation"),
        ("Scam Industries Ltd", "Unknown", "Multiple fraud reports"),
        ("Nicht Existiert AG", "Unknown", "Company does not exist"),
        ("Betrug & Partner KG", "Unknown", "Deceptive business practices"),
        ("Unseriös Consulting", "Consulting", "Unreliable, poor reviews"),
        ("Abzocke Services GmbH", "Services", "Reported for exploitation"),
        ("Phantomfirma SE", "Unknown", "Shell company"),
        ("Scheinfirma International", "Unknown", "No legitimate operations"),
    ]

    for row, (name, category, notes) in enumerate(blacklist_data, start=2):
        ws_blacklist.cell(row=row, column=1, value=name)
        ws_blacklist.cell(row=row, column=2, value=category)
        ws_blacklist.cell(row=row, column=3, value=notes)

    # Adjust column widths
    ws_blacklist.column_dimensions["A"].width = 35
    ws_blacklist.column_dimensions["B"].width = 20
    ws_blacklist.column_dimensions["C"].width = 50

    # Save workbook
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    print(f"Created sample Excel file: {output_file}")
    print(f"  Whitelist: {len(whitelist_data)} companies")
    print(f"  Blacklist: {len(blacklist_data)} companies")


if __name__ == "__main__":
    create_sample_excel()
