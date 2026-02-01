"""Excel file reader for company whitelist/blacklist."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from company_lookup.data.schemas import CompanyInfo, CompanyStatus, Config

logger = logging.getLogger(__name__)


class ExcelReader:
    """Reader for company lists from Excel files."""

    def __init__(self, config: Optional[Config] = None):
        """Initialize the Excel reader.

        Args:
            config: Configuration for column mappings and sheet names.
        """
        self.config = config or Config()
        self._companies: dict[str, CompanyInfo] = {}
        self._last_loaded: Optional[datetime] = None
        self._source_file: Optional[str] = None

    @property
    def companies(self) -> dict[str, CompanyInfo]:
        """Get all loaded companies."""
        return self._companies

    @property
    def whitelisted(self) -> list[CompanyInfo]:
        """Get all whitelisted companies."""
        return [c for c in self._companies.values() if c.status == CompanyStatus.WHITELISTED]

    @property
    def blacklisted(self) -> list[CompanyInfo]:
        """Get all blacklisted companies."""
        return [c for c in self._companies.values() if c.status == CompanyStatus.BLACKLISTED]

    def load_from_file(self, file_path: str) -> int:
        """Load company lists from an Excel file.

        Args:
            file_path: Path to the Excel file.

        Returns:
            Number of companies loaded.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file format is invalid.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if not path.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
            raise ValueError(f"Invalid file format: {path.suffix}. Expected .xlsx, .xls, or .xlsm")

        logger.info(f"Loading company lists from: {file_path}")

        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to open Excel file: {e}")

        self._companies.clear()
        loaded_count = 0

        # Load whitelist
        if self.config.whitelist_sheet in workbook.sheetnames:
            whitelist_sheet = workbook[self.config.whitelist_sheet]
            loaded_count += self._load_sheet(whitelist_sheet, CompanyStatus.WHITELISTED)
        else:
            logger.warning(f"Whitelist sheet '{self.config.whitelist_sheet}' not found")

        # Load blacklist
        if self.config.blacklist_sheet in workbook.sheetnames:
            blacklist_sheet = workbook[self.config.blacklist_sheet]
            loaded_count += self._load_sheet(blacklist_sheet, CompanyStatus.BLACKLISTED)
        else:
            logger.warning(f"Blacklist sheet '{self.config.blacklist_sheet}' not found")

        workbook.close()

        self._last_loaded = datetime.now()
        self._source_file = str(path.absolute())

        logger.info(f"Loaded {loaded_count} companies ({len(self.whitelisted)} whitelisted, {len(self.blacklisted)} blacklisted)")

        return loaded_count

    def _load_sheet(self, sheet: Worksheet, status: CompanyStatus) -> int:
        """Load companies from a single sheet.

        Args:
            sheet: The worksheet to load from.
            status: The status to assign to companies in this sheet.

        Returns:
            Number of companies loaded from this sheet.
        """
        loaded_count = 0
        headers: dict[str, int] = {}

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                # Parse headers
                headers = {
                    str(cell).strip().lower() if cell else "": idx
                    for idx, cell in enumerate(row)
                }
                continue

            # Get company name
            name_col = self._find_column_index(headers, self.config.company_name_column)
            if name_col is None or name_col >= len(row):
                continue

            company_name = row[name_col]
            if not company_name or not str(company_name).strip():
                continue

            company_name = str(company_name).strip()

            # Get optional fields
            notes = None
            if self.config.notes_column:
                notes_col = self._find_column_index(headers, self.config.notes_column)
                if notes_col is not None and notes_col < len(row) and row[notes_col]:
                    notes = str(row[notes_col]).strip()

            category = None
            if self.config.category_column:
                cat_col = self._find_column_index(headers, self.config.category_column)
                if cat_col is not None and cat_col < len(row) and row[cat_col]:
                    category = str(row[cat_col]).strip()

            # Create company info
            company = CompanyInfo(
                name=company_name,
                status=status,
                notes=notes,
                category=category,
                added_date=self._last_loaded,
            )

            # Use lowercase name as key for case-insensitive lookup
            key = company_name.lower()
            if key in self._companies:
                logger.warning(f"Duplicate company found: {company_name}")
            self._companies[key] = company
            loaded_count += 1

        return loaded_count

    def _find_column_index(self, headers: dict[str, int], column_name: str) -> Optional[int]:
        """Find the index of a column by name (case-insensitive).

        Args:
            headers: Dictionary mapping lowercase header names to indices.
            column_name: The column name to find.

        Returns:
            The column index or None if not found.
        """
        target = column_name.lower().strip()
        return headers.get(target)

    def get_company(self, name: str) -> Optional[CompanyInfo]:
        """Get a company by exact name match (case-insensitive).

        Args:
            name: The company name to look up.

        Returns:
            CompanyInfo if found, None otherwise.
        """
        return self._companies.get(name.lower().strip())

    def get_all_names(self) -> list[str]:
        """Get all company names.

        Returns:
            List of all company names.
        """
        return [c.name for c in self._companies.values()]

    def get_stats(self) -> dict:
        """Get statistics about the loaded companies.

        Returns:
            Dictionary with statistics.
        """
        categories = set()
        for company in self._companies.values():
            if company.category:
                categories.add(company.category)

        return {
            "total_companies": len(self._companies),
            "whitelisted_count": len(self.whitelisted),
            "blacklisted_count": len(self.blacklisted),
            "categories": sorted(categories),
            "last_updated": self._last_loaded,
            "source_file": self._source_file,
        }

    @staticmethod
    def create_template(file_path: str) -> None:
        """Create a template Excel file with the expected structure.

        Args:
            file_path: Path where to create the template file.
        """
        workbook = Workbook()

        # Create Whitelist sheet
        ws_whitelist = workbook.active
        ws_whitelist.title = "Whitelist"
        ws_whitelist.append(["Company Name", "Category", "Notes"])
        ws_whitelist.append(["Siemens AG", "Technology", "Major German corporation"])
        ws_whitelist.append(["BMW Group", "Automotive", "Car manufacturer"])
        ws_whitelist.append(["SAP SE", "Software", "Enterprise software"])

        # Create Blacklist sheet
        ws_blacklist = workbook.create_sheet("Blacklist")
        ws_blacklist.append(["Company Name", "Category", "Notes"])
        ws_blacklist.append(["Fake Company GmbH", "Unknown", "Known scam company"])
        ws_blacklist.append(["Scam Industries Ltd", "Unknown", "Fraudulent business"])

        # Adjust column widths
        for ws in [ws_whitelist, ws_blacklist]:
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 40

        workbook.save(file_path)
        logger.info(f"Created template Excel file: {file_path}")
